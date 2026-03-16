# -*- coding: utf-8 -*-
"""
@author:XuMing(xuming624@qq.com)
@description: Excel/Spreadsheet parser for TreeSearch.

Requires optional dependency: ``pip install openpyxl``
Extracts sheets, headers, and row data from Excel files and builds tree structure.
"""
import logging
import os
from typing import Optional

logger = logging.getLogger(__name__)

# Extensions supported by openpyxl
EXCEL_EXTENSIONS = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})


def _extract_excel_data(excel_path: str) -> list[dict]:
    """Extract sheet data from an Excel file.

    Returns a flat node list with:
    - Level 1: Sheet name
    - Level 2: Header row (columns)
    - Level 3: Data rows
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "Excel support requires 'openpyxl'. Install with: pip install openpyxl"
        )

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    nodes = []
    row_counter = 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            nodes.append({
                "title": sheet_name,
                "level": 1,
                "text": "(empty sheet)",
                "line_num": row_counter,
                "line_start": row_counter,
                "line_end": row_counter,
            })
            row_counter += 1
            continue

        # Detect header row (first row)
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        header_text = f"Columns: {', '.join(h for h in headers if h)}"

        # Sheet node (level 1)
        sheet_start = row_counter
        sheet_text_parts = [header_text]

        # Collect data rows
        data_rows_text = []
        for row_idx, row in enumerate(rows[1:], start=2):
            cells = [str(cell) if cell is not None else "" for cell in row]
            # Skip completely empty rows
            if not any(c.strip() for c in cells):
                continue
            row_text = "; ".join(
                f"{h}: {v}" for h, v in zip(headers, cells) if v.strip()
            )
            if row_text:
                data_rows_text.append(row_text)

        # Combine sheet content
        if data_rows_text:
            # Show up to 200 rows in text to avoid excessive node size
            displayed_rows = data_rows_text[:200]
            sheet_text_parts.extend(displayed_rows)
            if len(data_rows_text) > 200:
                sheet_text_parts.append(f"... ({len(data_rows_text) - 200} more rows)")

        sheet_text = "\n".join(sheet_text_parts)
        row_count = len(data_rows_text)

        nodes.append({
            "title": f"{sheet_name} ({row_count} rows)",
            "level": 1,
            "text": sheet_text,
            "line_num": sheet_start,
            "line_start": sheet_start,
            "line_end": sheet_start + row_count,
        })
        row_counter += row_count + 1

    wb.close()
    return nodes


async def excel_to_tree(
    excel_path: str,
    *,
    model: Optional[str] = None,
    if_add_node_summary: bool = True,
    summary_token_threshold: int = 200,
    if_add_doc_description: bool = False,
    if_add_node_text: bool = False,
    if_add_node_id: bool = True,
    **kwargs,
) -> dict:
    """Build a tree index from an Excel file.

    Each worksheet becomes a level-1 node. Headers and row data
    are extracted as structured text content.

    Returns:
        {'doc_name': str, 'structure': list, 'source_path': str}
    """
    doc_name = os.path.splitext(os.path.basename(excel_path))[0]
    logger.info("Parsing Excel: %s", excel_path)

    nodes = _extract_excel_data(excel_path)

    if not nodes:
        # Empty workbook, create a single root node
        nodes = [{"title": doc_name, "level": 1, "text": "(empty workbook)",
                  "line_num": 1, "line_start": 1, "line_end": 1}]

    from ..indexer import _build_tree, _finalize_tree

    tree = _build_tree(nodes)

    return _finalize_tree(
        tree, doc_name,
        source_path=os.path.abspath(excel_path),
        source_type="excel",
        if_add_node_id=if_add_node_id,
        if_add_node_summary=if_add_node_summary,
        summary_token_threshold=summary_token_threshold,
        if_add_node_text=if_add_node_text,
        if_add_doc_description=if_add_doc_description,
    )
