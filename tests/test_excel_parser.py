# -*- coding: utf-8 -*-
"""
@author:XuMing(xuming624@qq.com)
@description: Tests for the Excel parser.
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest

# Skip all tests if openpyxl is not available
openpyxl = pytest.importorskip("openpyxl")


@pytest.fixture
def sample_xlsx_file():
    """Create a temp Excel file with sample data."""
    from openpyxl import Workbook

    wb = Workbook()
    # First sheet: Sales data
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Product", "Revenue", "Quarter"])
    ws1.append(["Widget A", 10000, "Q1"])
    ws1.append(["Widget B", 25000, "Q2"])
    ws1.append(["Widget C", 15000, "Q3"])

    # Second sheet: Employees
    ws2 = wb.create_sheet("Employees")
    ws2.append(["Name", "Department", "Role"])
    ws2.append(["Alice", "Engineering", "Senior Engineer"])
    ws2.append(["Bob", "Marketing", "Marketing Manager"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    wb.save(path)
    wb.close()
    yield path
    os.unlink(path)


@pytest.fixture
def empty_xlsx_file():
    """Create a temp Excel file with an empty sheet."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "EmptySheet"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    wb.save(path)
    wb.close()
    yield path
    os.unlink(path)


async def test_excel_to_tree_basic(sample_xlsx_file):
    """Test basic Excel parsing produces correct tree structure."""
    from treesearch.parsers.excel_parser import excel_to_tree

    result = await excel_to_tree(sample_xlsx_file)

    assert "doc_name" in result
    assert "structure" in result
    assert "source_path" in result
    assert len(result["structure"]) == 2  # Two sheets

    # Check first sheet (Sales)
    sales_node = result["structure"][0]
    assert "Sales" in sales_node["title"]
    assert "3 rows" in sales_node["title"]
    assert "Widget A" in sales_node.get("text", sales_node.get("summary", ""))

    # Check second sheet (Employees)
    emp_node = result["structure"][1]
    assert "Employees" in emp_node["title"]
    assert "2 rows" in emp_node["title"]


async def test_excel_to_tree_empty(empty_xlsx_file):
    """Test parsing an Excel file with an empty sheet."""
    from treesearch.parsers.excel_parser import excel_to_tree

    result = await excel_to_tree(empty_xlsx_file)

    assert "structure" in result
    assert len(result["structure"]) >= 1


async def test_excel_to_tree_with_summary(sample_xlsx_file):
    """Test Excel parsing with summary generation enabled."""
    from treesearch.parsers.excel_parser import excel_to_tree

    result = await excel_to_tree(
        sample_xlsx_file,
        if_add_node_summary=True,
        if_add_node_text=True,
    )

    assert "structure" in result
    # Each node should have a summary or text
    for node in result["structure"]:
        has_content = (
            node.get("summary")
            or node.get("prefix_summary")
            or node.get("text")
        )
        assert has_content, f"Node '{node.get('title')}' has no content"


async def test_excel_to_tree_with_node_id(sample_xlsx_file):
    """Test that node IDs are assigned."""
    from treesearch.parsers.excel_parser import excel_to_tree

    result = await excel_to_tree(sample_xlsx_file, if_add_node_id=True)

    for node in result["structure"]:
        assert "node_id" in node


async def test_excel_parser_registered():
    """Test that Excel parser is registered in ParserRegistry."""
    from treesearch.parsers.registry import ParserRegistry, SOURCE_TYPE_MAP

    parser = ParserRegistry.get(".xlsx")
    assert parser is not None, "Excel parser not registered for .xlsx"

    assert SOURCE_TYPE_MAP.get(".xlsx") == "excel"
    assert SOURCE_TYPE_MAP.get(".xlsm") == "excel"


async def test_excel_integration_with_treesearch(sample_xlsx_file):
    """Test Excel file works with TreeSearch end-to-end."""
    from treesearch import TreeSearch

    ts = TreeSearch(sample_xlsx_file, db_path=None)
    results = await ts.asearch("Widget Revenue")

    assert "documents" in results
    assert len(results["documents"]) > 0
